# Copyright (c) 2025 Bytedance Ltd. and/or its affiliates
# SPDX-License-Identifier: MIT

"""把结构化表格数据渲染成 xlsx。纯 openpyxl，无外部依赖."""

import os
import uuid

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
_HEADER_FONT = Font(bold=True)
_THIN = Side(style="thin", color="999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def build_xlsx(data: dict) -> str:
    """表格数据 -> xlsx 文件路径。

    data: {title, sheets:[{name, headers:[str], rows:[[…]], note?}]}
    样式：表头加粗+浅蓝底、全表细边框、列宽按内容自适应（8~40 字符）。
    """
    wb = Workbook()
    wb.remove(wb.active)
    sheets = data.get("sheets") or [{"name": "Sheet1", "headers": [], "rows": []}]
    for idx, sheet in enumerate(sheets):
        name = str(sheet.get("name") or f"Sheet{idx + 1}")[:31]
        ws = wb.create_sheet(title=name)
        headers = [str(h) for h in (sheet.get("headers") or [])]
        rows = sheet.get("rows") or []
        if headers:
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.border = _BORDER
                cell.alignment = Alignment(horizontal="center")
        for row in rows:
            cells = ["" if v is None else v for v in row]
            ws.append(cells)
            for cell in ws[ws.max_row]:
                cell.border = _BORDER
        # 列宽自适应：中文按 2 字符估宽，夹在 8~40
        col_count = max(len(headers), *(len(r) for r in rows)) if rows or headers else 0
        for col in range(1, col_count + 1):
            width = 8
            for row_cells in ws.iter_rows(min_col=col, max_col=col, values_only=True):
                v = row_cells[0]
                if v is None:
                    continue
                text = str(v)
                est = sum(2 if ord(ch) > 0x2E80 else 1 for ch in text)
                width = max(width, est + 2)
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(
                width, 40
            )
        note = sheet.get("note")
        if note:
            ws.append([])
            ws.append([f"说明：{note}"])
    path = os.path.join(os.getcwd(), f"sheet_{uuid.uuid4().hex}.xlsx")
    wb.save(path)
    return path
