# Copyright (c) 2025 Bytedance Ltd. and/or its affiliates
# SPDX-License-Identifier: MIT

"""办公文档（doc）与数据表格（sheet）技能的导出器/预设单元测试."""

import os

from src.skills.presets import SUB_SKILL_PRESETS, resolve_sub_skill


class TestOfficePresets:
    def test_doc_presets_complete(self):
        assert set(SUB_SKILL_PRESETS["doc"]) == {
            "weekly",
            "minutes",
            "plan",
            "notice",
            "resume",
        }
        # 办公文档没有「通用」默认：每个子能力都有实质预设
        for sub, text in SUB_SKILL_PRESETS["doc"].items():
            assert text, f"doc:{sub} 预设不能为空"

    def test_sheet_presets_complete(self):
        assert set(SUB_SKILL_PRESETS["sheet"]) == {
            "general",
            "timetable",
            "duty",
            "budget",
            "tracker",
        }
        assert resolve_sub_skill("sheet", "general") == ("general", "")

    def test_unknown_still_falls_back(self):
        assert resolve_sub_skill("doc", "nope") == (None, "")
        assert resolve_sub_skill("sheet", None) == (None, "")


class TestSectionsDocx:
    def test_build_sections_docx(self):
        from src.skills.docx_export import build_sections_docx

        path = build_sections_docx(
            {
                "title": "第 29 周工作周报",
                "meta": "运营部 · 2026-07-20",
                "sections": [
                    {"heading": "一、本周完成", "content": "上线A\n推进B"},
                    {"heading": "二、下周计划", "content": "启动C"},
                ],
            },
            default_title="文档",
            prefix="document",
        )
        assert os.path.exists(path) and os.path.getsize(path) > 10_000
        import docx

        texts = [p.text for p in docx.Document(path).paragraphs]
        assert "第 29 周工作周报" in texts
        assert "上线A" in texts and "推进B" in texts  # 换行拆段
        os.remove(path)

    def test_speech_still_works(self):
        from src.skills.docx_export import build_speech_docx

        path = build_speech_docx(
            {"title": "《春》说课稿", "sections": [{"heading": "一、说教材", "content": "略"}]}
        )
        assert os.path.exists(path)
        os.remove(path)


class TestXlsxBuilder:
    def test_build_xlsx_roundtrip(self):
        from src.skills.xlsx_export import build_xlsx

        path = build_xlsx(
            {
                "title": "班级课程表",
                "sheets": [
                    {
                        "name": "初一(3)班",
                        "headers": ["节次", "周一", "周二"],
                        "rows": [
                            ["第1节 8:00-8:45", "语文", "数学"],
                            ["第2节 8:55-9:40", "数学", "英语"],
                        ],
                        "note": "每周一升旗仪式后第一节顺延 10 分钟",
                    }
                ],
            }
        )
        assert os.path.exists(path)
        import openpyxl

        wb = openpyxl.load_workbook(path)
        ws = wb["初一(3)班"]
        assert ws.cell(1, 1).value == "节次"
        assert ws.cell(1, 1).font.bold
        assert ws.cell(2, 2).value == "语文"
        assert ws.max_row == 5  # 2 数据行 + 表头 + 空行 + 说明
        assert "说明：" in str(ws.cell(5, 1).value)
        os.remove(path)

    def test_build_xlsx_empty_safe(self):
        from src.skills.xlsx_export import build_xlsx

        path = build_xlsx({"title": "空表"})
        assert os.path.exists(path)
        os.remove(path)

    def test_sheet_name_truncated(self):
        from src.skills.xlsx_export import build_xlsx

        path = build_xlsx(
            {"sheets": [{"name": "超" * 40, "headers": ["a"], "rows": [["1"]]}]}
        )
        import openpyxl

        wb = openpyxl.load_workbook(path)
        assert len(wb.sheetnames[0]) == 31
        os.remove(path)
